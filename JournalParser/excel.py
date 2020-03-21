from openpyxl import Workbook
from openpyxl.styles import Font



def excelify(warnings: dict, order: str = 'grades'):
    file = Workbook()

    if order == 'grades':
        for grade, subjs in warnings.items():
            sheet = file.create_sheet(grade) #name sheet as a grade
            for subj_name, warns_list in subjs.items():
                sheet.append([subj_name]) # for each subj write its name
                makebold(sheet)
                sheet.append([warns_list[0].teacher]) #and teacher
                makebold(sheet)
                for w_list in warns_list:
                    sheet.append([f'Четверть {w_list.term}']) #write term
                    for row in w_list.warnings:
                        sheet.append(row)
                    create_space(sheet) #gap between terms
                create_space(sheet) #gap between subjects

    else: #by teachers ordering
        for grade, subjs in warnings.items():
            for subj_name, warns_list in subjs.items():
                teacher = warns_list[0].teacher
                initials = teacher.split()
                initials = ' '.join([initials[0], initials[1][0], initials[-1][0]])

                if initials in file.sheetnames:
                    sheet = file.get_sheet_by_name(initials)
                else:
                    sheet = file.create_sheet(initials)

                for w_list in warns_list:
                    sheet.append([subj_name])
                    makebold(sheet)
                    sheet.append([grade])
                    makebold(sheet)
                    sheet.append([f'Четверть {w_list.term}'])
                    for row in w_list.warnings:
                        sheet.append(row)
                    create_space(sheet)
                create_space(sheet)

    file.save('Проверка журналов.xlsx')

def create_space(sheet, n=1):
    for i in range(n):
        sheet.append([])

def makebold(sheet):
    bold = Font(bold=True)
    cell = sheet.cell(sheet.max_row, 1)
    cell.font = bold