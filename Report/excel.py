from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment
import os
from .objects import OverallQualitiesTable, SubjectsQualitiesTable, StudentsPerformanceTables


def excelify_oqt(**kwargs):
    file = Workbook()

    for key, val in kwargs.items():
        sheet = file.create_sheet(key)
        for table in val:
            sheet.append(table.header)
            for row in table.data:
                sheet.append(row)
            create_space(sheet)
            if isinstance(table, OverallQualitiesTable): # Create chart only on the first page
                create_chart(sheet, len(table.data[0]) - 2, len(table.data))# exlude dynamics and grade cols

    filename = 'Результативность работы школы за учебный период.xlsx'

    file.save(filename)

def excelify_sqt(**kwargs):

    file = Workbook()

    for key, val in kwargs.items():
        sheet = file.create_sheet(key)
        for table in val:
            if isinstance(table, SubjectsQualitiesTable):
                data = table.data
                for subj in data.keys():
                    sheet.append([subj]) #Subject
                    sheet.append(table.header) #table header
                    for grade_row in data[subj]:
                        sheet.append(grade_row) #data
                    create_chart(sheet, len(table.header) - 2, len(data[subj]))
                    create_space(sheet, n=8)
            else:
                sheet.append(table.header)
                for row in table.data:
                    sheet.append(row)
                create_space(sheet)

    filename = 'Результативность работы школы (по предметам) за учебный период.xlsx'

    file.save(filename)


def excelify_spt(tables: 'StudentsPerformanceTables'):
    file = Workbook()
    align = Alignment(wrap_text=True, shrinkToFit=True)

    chart_data_width = len(tables.header_for_count) - 2
    chart_data_heigth = len(tables.one_fours_count.keys())

    sheet = file.create_sheet('Кол-во отличников')
    sheet.append(tables.header_for_exc)
    for grade, data in tables.excellents.items():
        sheet.append([grade, *data])

    sheet = file.create_sheet('С одной 4')
    sheet.append(tables.header_for_list)
    for row in tables.one_fours:
        sheet.append(row)
    create_space(sheet)
    sheet.append(tables.header_for_count)
    for grade, data in tables.one_fours_count.items():
        sheet.append([grade, *data])
    create_space(sheet)
    create_chart(sheet, chart_data_width, chart_data_heigth)

    sheet = file.create_sheet('С двумя 4')
    sheet.append(tables.header_for_list)
    for row in tables.two_fours:
        sheet.append(row)
    create_space(sheet)
    sheet.append(tables.header_for_count)
    for grade, data in tables.two_fours_count.items():
        sheet.append([grade, *data])
    create_space(sheet)
    create_chart(sheet, chart_data_width, chart_data_heigth)

    sheet = file.create_sheet('С одной 3')
    sheet.append(tables.header_for_list)
    for row in tables.one_threes:
        sheet.append(row)
    create_space(sheet)
    sheet.append(tables.header_for_count)
    for grade, data in tables.one_threes_count.items():
        sheet.append([grade, *data])
    create_space(sheet)
    create_chart(sheet, chart_data_width, chart_data_heigth)

    sheet = file.create_sheet('С двумя 3')
    sheet.append(tables.header_for_list)
    for row in tables.two_threes:
        sheet.append(row)
    create_space(sheet)
    sheet.append(tables.header_for_count)
    for grade, data in tables.two_threes_count.items():
        sheet.append([grade, *data])
    create_space(sheet)
    create_chart(sheet, chart_data_width, chart_data_heigth)

    for sheet in file.worksheets[1:]:
        for row in sheet.rows:
            for cell in row:
                cell.alignment = align

    filename = 'Итоги успеваемости класса за учебный период.xlsx'

    file.save(filename)

def excelify_bst(bad_stud_list: list):
    file = Workbook()
    sheet = file.create_sheet('Ученики')

    for row in bad_stud_list:
        sheet.append(row)

    filename = 'Ср. балл ниже 3.xlsx'

    file.save(filename)

def create_space(sheet, n=2):
    for i in range(n):
        sheet.append([])


def create_chart(sheet, width, height):

    chart = BarChart()
    chart.type = 'col'
    chart.style = 10
    chart.y_axis.title = 'Качество, %'
    chart.x_axis.title = 'Классы'

    data = Reference(sheet, min_col=2, max_col=width + 1, min_row=sheet.max_row - height, max_row=sheet.max_row)
    cats = Reference(sheet, min_col=1, min_row=sheet.max_row - height + 1, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    sheet.add_chart(chart, 'K{}'.format(sheet.max_row - height))

